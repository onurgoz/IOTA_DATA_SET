#kütüphaneleri yüklüyoruz
from bs4 import BeautifulSoup
import requests
from  openpyxl import *

#Veri çekme için gerekli parametreler
headers_param = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}
#Verileri çekeceğim web sayfası
Link="https://www.coinlore.com/coin/iota/historical-data/-7200/1605301199#hist-prices"

#Dosya işlemleri
Data_xlsx = Workbook()
sayfa1=Data_xlsx.create_sheet("İslenmemis Veri",0)
sayfa2=Data_xlsx.create_sheet("İslenmis Veri",1)


#Veriyi parselleme
IOTA_page=requests.get(f"{Link}",headers=headers_param)
soup = BeautifulSoup(IOTA_page.content,"html.parser")

#Veriyi ayıklama
ozellikler= list(soup.find_all("td",attrs={"class":"no-wrap"}))

#İşlenmemiş verimizi kaydediyoruz
for i in range(0,len(ozellikler)):
  sayfa1.append([ozellikler[i].text])

#Verideki $ işaretlerini temizliyoruz
for i in range(0,len(ozellikler)):
  ozellikler[i]=ozellikler[i].text.replace("$","")

#Elimizdeki veriyi parçalayıp xlsx dosyasına yazıyoruz

sayfa2.append(['Date','Open','High','Low','Close','Volume','Volume IOTA','Market Cap'])

for i in range(0,9863,8):
   sayfa2.append([ozellikler[0+i],ozellikler[1+i],ozellikler[2+i],ozellikler[3+i],ozellikler[4+i],ozellikler[5+i],ozellikler[6+i],ozellikler[7+i]])

#Dosyayı kaydedip kapatıyoruz
Data_xlsx.save("IOTA_DATA.xlsx")
Data_xlsx.close()